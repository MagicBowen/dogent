from __future__ import annotations

import asyncio
import base64
import io
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import unquote, urlparse
from xml.etree import ElementTree as ET

from ..config.resources import read_config_text
from ..core.session_log import log_exception

DEFAULT_MAX_CHARS = 15000
DEFAULT_XLSX_MAX_ROWS = 50
DEFAULT_XLSX_MAX_COLS = 20
PACKAGE_MODE_ENV = "DOGENT_PACKAGE_MODE"
PDF_STYLE_FILENAME = "pdf_style.css"


@dataclass(frozen=True)
class DocumentReadResult:
    content: str
    truncated: bool
    format: str
    metadata: dict[str, Any]
    error: str | None = None


@dataclass(frozen=True)
class DocumentConvertResult:
    input_format: str
    output_format: str
    output_path: Path
    extracted_media_dir: Path | None
    notes: list[str]


def _package_mode() -> str:
    mode = os.getenv(PACKAGE_MODE_ENV, "lite").strip().lower()
    return "full" if mode == "full" else "lite"


def _platform_tag() -> str:
    if sys.platform == "win32":
        return "windows"
    if sys.platform == "darwin":
        return "macos"
    return "linux"


def _bundled_tools_root() -> Path:
    return Path(__file__).resolve().parents[1] / "resources" / "tools"


def _resolve_pandoc_binary() -> Path | None:
    if _package_mode() != "full":
        return None
    root = _bundled_tools_root() / "pandoc" / _platform_tag()
    binary_name = "pandoc.exe" if sys.platform == "win32" else "pandoc"
    candidate = root / binary_name
    if candidate.is_file():
        return candidate
    return None


def _resolve_playwright_browsers_path() -> Path | None:
    if _package_mode() != "full":
        return None
    root = _bundled_tools_root() / "playwright" / _platform_tag()
    if not root.exists():
        return None
    for child in root.iterdir():
        if child.is_dir() and child.name.startswith("chromium"):
            return root
    return None


def read_document(
    path: Path,
    *,
    sheet: str | None = None,
    max_chars: int = DEFAULT_MAX_CHARS,
    offset: int = 0,
    length: int | None = None,
) -> DocumentReadResult:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return _read_pdf(path, max_chars=max_chars, offset=offset, length=length)
    if ext == ".docx":
        return _read_docx(path, max_chars=max_chars, offset=offset, length=length)
    if ext == ".xlsx":
        return _read_xlsx(
            path,
            sheet=sheet,
            max_chars=max_chars,
            offset=offset,
            length=length,
        )
    return _read_text(path, max_chars=max_chars, offset=offset, length=length)


def export_markdown(
    md_path: Path,
    *,
    output_path: Path,
    format: str,
    title: str | None = None,
    workspace_root: Path | None = None,
) -> list[str]:
    normalized = format.strip().lower()
    if normalized == "docx":
        _markdown_to_docx(md_path, output_path=output_path, workspace_root=workspace_root)
        return []
    if normalized == "pdf":
        return _run_async(
            _markdown_to_pdf(
                md_path,
                output_path=output_path,
                title=title,
                workspace_root=workspace_root,
            )
        )
    raise ValueError(f"Unsupported export format: {format}")


async def export_markdown_async(
    md_path: Path,
    *,
    output_path: Path,
    format: str,
    title: str | None = None,
    workspace_root: Path | None = None,
) -> list[str]:
    normalized = format.strip().lower()
    if normalized == "docx":
        _markdown_to_docx(md_path, output_path=output_path, workspace_root=workspace_root)
        return []
    if normalized == "pdf":
        return await _markdown_to_pdf(
            md_path,
            output_path=output_path,
            title=title,
            workspace_root=workspace_root,
        )
    raise ValueError(f"Unsupported export format: {format}")


async def convert_document_async(
    input_path: Path,
    *,
    output_path: Path,
    extract_media_dir: Path | None = None,
    workspace_root: Path | None = None,
) -> DocumentConvertResult:
    input_format = _detect_format(input_path)
    output_format = _detect_format(output_path)
    if not input_format:
        raise ValueError("Unsupported input format. Use docx, pdf, md, or xlsx.")
    if not output_format:
        raise ValueError("Unsupported output format. Use docx, pdf, or md.")
    if input_format == output_format:
        raise ValueError("Input and output formats are the same.")
    if extract_media_dir and not (
        input_format == "docx" and output_format == "md"
    ):
        raise ValueError("extract_media_dir is only supported for DOCX to Markdown.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    notes: list[str] = []

    if input_format == "docx" and output_format == "md":
        _docx_to_markdown(
            input_path,
            output_path=output_path,
            extract_media_dir=extract_media_dir,
        )
    elif input_format == "md" and output_format == "docx":
        _markdown_to_docx(input_path, output_path=output_path, workspace_root=workspace_root)
    elif input_format == "md" and output_format == "pdf":
        style_notes = await _markdown_to_pdf(
            input_path,
            output_path=output_path,
            title=output_path.stem,
            workspace_root=workspace_root,
        )
        notes.extend(style_notes)
    elif input_format == "docx" and output_format == "pdf":
        notes.append("Converted DOCX -> Markdown -> PDF; formatting may differ.")
        with tempfile.TemporaryDirectory() as tmp:
            tmp_md = Path(tmp) / "source.md"
            _docx_to_markdown(input_path, output_path=tmp_md, extract_media_dir=None)
            style_notes = await _markdown_to_pdf(
                tmp_md,
                output_path=output_path,
                title=output_path.stem,
                workspace_root=workspace_root,
            )
            notes.extend(style_notes)
    elif input_format == "pdf" and output_format == "md":
        notes.append("PDF text extracted; layout and images are not preserved.")
        result = read_document(input_path, max_chars=0)
        if result.error:
            raise RuntimeError(result.error)
        output_path.write_text(result.content, encoding="utf-8")
    elif input_format == "pdf" and output_format == "docx":
        notes.append("PDF text extracted; layout and images are not preserved.")
        result = read_document(input_path, max_chars=0)
        if result.error:
            raise RuntimeError(result.error)
        with tempfile.TemporaryDirectory() as tmp:
            tmp_md = Path(tmp) / "source.md"
            tmp_md.write_text(result.content, encoding="utf-8")
            _markdown_to_docx(tmp_md, output_path=output_path, workspace_root=workspace_root)
    elif input_format == "xlsx" and output_format == "md":
        result = read_document(input_path, max_chars=0)
        if result.error:
            raise RuntimeError(result.error)
        output_path.write_text(result.content, encoding="utf-8")
    else:
        raise ValueError(f"Unsupported conversion: {input_format} -> {output_format}")

    return DocumentConvertResult(
        input_format=input_format,
        output_format=output_format,
        output_path=output_path,
        extracted_media_dir=extract_media_dir,
        notes=notes,
    )


def _read_text(
    path: Path,
    *,
    max_chars: int,
    offset: int,
    length: int | None,
) -> DocumentReadResult:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        return DocumentReadResult(
            content="",
            truncated=False,
            format=_format_from_suffix(path),
            metadata={},
            error=str(exc),
        )
    content, truncated, paging = _apply_size_limit(
        text,
        max_chars,
        offset=offset,
        length=length,
    )
    return DocumentReadResult(
        content=content,
        truncated=truncated,
        format=_format_from_suffix(path),
        metadata=paging,
    )


def _read_docx(
    path: Path,
    *,
    max_chars: int,
    offset: int,
    length: int | None,
) -> DocumentReadResult:
    try:
        _ensure_pandoc_available()
        import pypandoc

        text = pypandoc.convert_file(
            str(path),
            to="markdown",
            format="docx",
            extra_args=["--track-changes=all"],
        )
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        return DocumentReadResult(
            content="",
            truncated=False,
            format="docx",
            metadata={},
            error=f"DOCX read failed: {exc}",
        )
    content, truncated, paging = _apply_size_limit(
        text,
        max_chars,
        offset=offset,
        length=length,
    )
    return DocumentReadResult(
        content=content,
        truncated=truncated,
        format="docx",
        metadata=paging,
    )


def _docx_to_markdown(
    input_path: Path,
    *,
    output_path: Path,
    extract_media_dir: Path | None,
) -> None:
    _ensure_pandoc_available()
    import pypandoc

    extra_args = ["--track-changes=all"]
    if extract_media_dir:
        extract_media_dir.mkdir(parents=True, exist_ok=True)
        extra_args.append(f"--extract-media={extract_media_dir}")
    pypandoc.convert_file(
        str(input_path),
        to="markdown",
        format="docx",
        outputfile=str(output_path),
        extra_args=extra_args,
    )


def _read_pdf(
    path: Path,
    *,
    max_chars: int,
    offset: int,
    length: int | None,
) -> DocumentReadResult:
    try:
        import fitz
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        return DocumentReadResult(
            content="",
            truncated=False,
            format="pdf",
            metadata={},
            error=f"PDF read failed (missing PyMuPDF): {exc}",
        )

    try:
        doc = fitz.open(str(path))
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        return DocumentReadResult(
            content="",
            truncated=False,
            format="pdf",
            metadata={},
            error=f"PDF read failed: {exc}",
        )

    try:
        parts: list[str] = []
        pages_with_text = 0
        for idx, page in enumerate(doc, start=1):
            text = page.get_text("text").strip()
            if text:
                pages_with_text += 1
                parts.append(f"<!-- page:{idx} -->\n\n{text}\n")
        metadata = {"pages": len(doc)}
        if pages_with_text == 0:
            return DocumentReadResult(
                content="",
                truncated=False,
                format="pdf",
                metadata=metadata,
                error="Unsupported PDF: no extractable text (scanned PDF not supported).",
            )
        combined = "\n\n".join(parts).strip() + "\n"
        content, truncated, paging = _apply_size_limit(
            combined,
            max_chars,
            offset=offset,
            length=length,
        )
        return DocumentReadResult(
            content=content,
            truncated=truncated,
            format="pdf",
            metadata={**metadata, **paging},
        )
    finally:
        doc.close()


def _read_xlsx(
    path: Path,
    *,
    sheet: str | None,
    max_chars: int,
    offset: int,
    length: int | None,
) -> DocumentReadResult:
    try:
        import openpyxl
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        return DocumentReadResult(
            content="",
            truncated=False,
            format="xlsx",
            metadata={},
            error=f"XLSX read failed (missing openpyxl): {exc}",
        )
    try:
        workbook = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        return _read_xlsx_xml(
            path,
            sheet=sheet,
            max_chars=max_chars,
            offset=offset,
            length=length,
            error_context=exc,
        )
    try:
        sheetnames = workbook.sheetnames
        if sheet:
            sheet_name = _select_sheet(sheetnames, sheet)
            if not sheet_name:
                return DocumentReadResult(
                    content="",
                    truncated=False,
                    format="xlsx",
                    metadata={"sheets": sheetnames},
                    error=f"Sheet not found: {sheet}",
                )
            ws = workbook[sheet_name]
            total_rows = int(ws.max_row or 0)
            total_cols = int(ws.max_column or 0)
            if total_rows == 0 or total_cols == 0:
                table, meta = _sheet_to_markdown_table(
                    [],
                    total_rows=total_rows,
                    total_cols=total_cols,
                    max_rows=DEFAULT_XLSX_MAX_ROWS,
                    max_cols=DEFAULT_XLSX_MAX_COLS,
                )
            else:
                capped_rows = min(total_rows, DEFAULT_XLSX_MAX_ROWS)
                capped_cols = min(total_cols, DEFAULT_XLSX_MAX_COLS)
                table, meta = _sheet_to_markdown_table(
                    ws.iter_rows(
                        min_row=1,
                        max_row=capped_rows,
                        max_col=capped_cols,
                        values_only=True,
                    ),
                    total_rows=total_rows,
                    total_cols=total_cols,
                    max_rows=DEFAULT_XLSX_MAX_ROWS,
                    max_cols=DEFAULT_XLSX_MAX_COLS,
                )
            metadata = {"sheet": sheet_name, **meta}
            content, truncated, paging = _apply_size_limit(
                table,
                max_chars,
                offset=offset,
                length=length,
            )
            return DocumentReadResult(
                content=content,
                truncated=truncated,
                format="xlsx",
                metadata={**metadata, **paging},
            )
        if not sheetnames:
            return DocumentReadResult(
                content="",
                truncated=False,
                format="xlsx",
                metadata={},
                error="No sheets found in XLSX.",
            )
        sections = [f"# {path.stem}"]
        sheets_meta: list[dict[str, Any]] = []
        for sheet_name in sheetnames:
            ws = workbook[sheet_name]
            total_rows = int(ws.max_row or 0)
            total_cols = int(ws.max_column or 0)
            if total_rows == 0 or total_cols == 0:
                table, meta = _sheet_to_markdown_table(
                    [],
                    total_rows=total_rows,
                    total_cols=total_cols,
                    max_rows=DEFAULT_XLSX_MAX_ROWS,
                    max_cols=DEFAULT_XLSX_MAX_COLS,
                )
            else:
                capped_rows = min(total_rows, DEFAULT_XLSX_MAX_ROWS)
                capped_cols = min(total_cols, DEFAULT_XLSX_MAX_COLS)
                table, meta = _sheet_to_markdown_table(
                    ws.iter_rows(
                        min_row=1,
                        max_row=capped_rows,
                        max_col=capped_cols,
                        values_only=True,
                    ),
                    total_rows=total_rows,
                    total_cols=total_cols,
                    max_rows=DEFAULT_XLSX_MAX_ROWS,
                    max_cols=DEFAULT_XLSX_MAX_COLS,
                )
            sheets_meta.append({"name": sheet_name, **meta})
            sections.append(f"## {sheet_name}")
            sections.append(table)
        joined = "\n\n".join(sections)
        content, truncated, paging = _apply_size_limit(
            joined,
            max_chars,
            offset=offset,
            length=length,
        )
        return DocumentReadResult(
            content=content,
            truncated=truncated,
            format="xlsx",
            metadata={"sheets": sheetnames, "sheets_meta": sheets_meta, **paging},
        )
    finally:
        try:
            workbook.close()
        except Exception as exc:
            log_exception("document_io", exc)
            pass


def _read_xlsx_xml(
    path: Path,
    *,
    sheet: str | None,
    max_chars: int,
    offset: int,
    length: int | None,
    error_context: Exception | None = None,
) -> DocumentReadResult:
    try:
        with zipfile.ZipFile(path) as zf:
            sheet_info = _xlsx_sheet_info(zf)
            if not sheet_info:
                return DocumentReadResult(
                    content="",
                    truncated=False,
                    format="xlsx",
                    metadata={},
                    error="No sheets found in XLSX.",
                )
            sheetnames = [name for name, _ in sheet_info]
            shared_strings = _xlsx_shared_strings(zf)
            if sheet:
                sheet_name, sheet_path = _xlsx_select_sheet(sheet_info, sheet)
                if not sheet_name:
                    return DocumentReadResult(
                        content="",
                        truncated=False,
                        format="xlsx",
                        metadata={"sheets": sheetnames},
                        error=f"Sheet not found: {sheet}",
                    )
                table, meta = _xlsx_sheet_to_markdown(
                    zf,
                    sheet_path,
                    shared_strings,
                    max_rows=DEFAULT_XLSX_MAX_ROWS,
                    max_cols=DEFAULT_XLSX_MAX_COLS,
                )
                metadata = {"sheet": sheet_name, **meta}
                content, truncated, paging = _apply_size_limit(
                    table,
                    max_chars,
                    offset=offset,
                    length=length,
                )
                return DocumentReadResult(
                    content=content,
                    truncated=truncated,
                    format="xlsx",
                    metadata={**metadata, **paging},
                )
            sections = [f"# {path.stem}"]
            sheets_meta: list[dict[str, Any]] = []
            for sheet_name, sheet_path in sheet_info:
                table, meta = _xlsx_sheet_to_markdown(
                    zf,
                    sheet_path,
                    shared_strings,
                    max_rows=DEFAULT_XLSX_MAX_ROWS,
                    max_cols=DEFAULT_XLSX_MAX_COLS,
                )
                sheets_meta.append({"name": sheet_name, **meta})
                sections.append(f"## {sheet_name}")
                sections.append(table)
            joined = "\n\n".join(sections)
            content, truncated, paging = _apply_size_limit(
                joined,
                max_chars,
                offset=offset,
                length=length,
            )
            return DocumentReadResult(
                content=content,
                truncated=truncated,
                format="xlsx",
                metadata={"sheets": sheetnames, "sheets_meta": sheets_meta, **paging},
            )
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        hint = f"{error_context}; " if error_context else ""
        return DocumentReadResult(
            content="",
            truncated=False,
            format="xlsx",
            metadata={},
            error=f"XLSX read failed: {hint}{exc}",
        )


def _sheet_to_markdown_table(
    rows: Iterable[Iterable[object | None]],
    *,
    total_rows: int,
    total_cols: int,
    max_rows: int,
    max_cols: int,
) -> tuple[str, dict[str, Any]]:
    if total_rows == 0 or total_cols == 0:
        return "(empty sheet)", {
            "rows": total_rows,
            "cols": total_cols,
            "truncated": False,
        }

    capped_rows = min(total_rows, max_rows)
    capped_cols = min(total_cols, max_cols)
    raw_rows = [list(row) for row in rows]
    if not raw_rows:
        return "(empty sheet)", {
            "rows": total_rows,
            "cols": total_cols,
            "truncated": False,
        }
    header = raw_rows[0][:capped_cols]
    body_rows = raw_rows[1:capped_rows]

    if not any(_cell_to_str(cell) for cell in header):
        header = [_excel_col_name(idx + 1) for idx in range(capped_cols)]

    lines = []
    lines.append("| " + " | ".join(_cell_to_str(cell) for cell in header) + " |")
    lines.append("| " + " | ".join(["---"] * capped_cols) + " |")
    for row in body_rows:
        padded = list(row[:capped_cols])
        if len(padded) < capped_cols:
            padded.extend([None] * (capped_cols - len(padded)))
        lines.append("| " + " | ".join(_cell_to_str(cell) for cell in padded) + " |")

    truncated = total_rows > capped_rows or total_cols > capped_cols
    if truncated:
        lines.append(
            f"\n[Truncated to {capped_rows} rows x {capped_cols} cols]"
        )
    metadata = {"rows": total_rows, "cols": total_cols, "truncated": truncated}
    return "\n".join(lines), metadata


def _cell_to_str(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return str(value)


def _excel_col_name(index: int) -> str:
    name = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


def _excel_col_index(label: str) -> int:
    total = 0
    for ch in label.upper():
        if not ("A" <= ch <= "Z"):
            break
        total = total * 26 + (ord(ch) - ord("A") + 1)
    return total


def _select_sheet(sheetnames: list[str], requested: str | None) -> str | None:
    if not sheetnames:
        return None
    if requested is None or not requested.strip():
        return sheetnames[0]
    requested = requested.strip()
    if requested in sheetnames:
        return requested
    lowered = requested.lower()
    for name in sheetnames:
        if name.lower() == lowered:
            return name
    return None


def _xlsx_strip_ns(tag: str) -> str:
    return tag.split("}", 1)[-1]


def _xlsx_sheet_info(zf: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook_xml = zf.read("xl/workbook.xml")
    rels_xml = zf.read("xl/_rels/workbook.xml.rels")
    workbook_root = ET.fromstring(workbook_xml)
    rels_root = ET.fromstring(rels_xml)
    rel_map: dict[str, str] = {}
    for rel in rels_root.iter():
        if _xlsx_strip_ns(rel.tag) != "Relationship":
            continue
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rel_map[rel_id] = target
    sheet_info: list[tuple[str, str]] = []
    for sheet in workbook_root.iter():
        if _xlsx_strip_ns(sheet.tag) != "sheet":
            continue
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        ) or sheet.attrib.get("r:id")
        target = rel_map.get(rel_id or "")
        if not (name and target):
            continue
        if target.startswith("/"):
            target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        sheet_info.append((name, target))
    return sheet_info


def _xlsx_select_sheet(
    sheet_info: list[tuple[str, str]], requested: str
) -> tuple[str | None, str | None]:
    requested = requested.strip()
    for name, path in sheet_info:
        if name == requested:
            return name, path
    lowered = requested.lower()
    for name, path in sheet_info:
        if name.lower() == lowered:
            return name, path
    return None, None


def _xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        data = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ET.fromstring(data)
    strings: list[str] = []
    for si in root.iter():
        if _xlsx_strip_ns(si.tag) != "si":
            continue
        parts: list[str] = []
        for node in si.iter():
            if _xlsx_strip_ns(node.tag) == "t":
                parts.append(node.text or "")
        strings.append("".join(parts))
    return strings


def _xlsx_cell_value(
    cell: ET.Element, shared_strings: list[str]
) -> object | None:
    cell_type = cell.attrib.get("t")
    v_text = None
    inline_parts: list[str] = []
    for child in cell:
        tag = _xlsx_strip_ns(child.tag)
        if tag == "v":
            v_text = child.text
        elif tag == "is":
            for node in child.iter():
                if _xlsx_strip_ns(node.tag) == "t":
                    inline_parts.append(node.text or "")
    if cell_type == "inlineStr":
        return "".join(inline_parts)
    if cell_type == "s":
        if v_text is None:
            return ""
        try:
            idx = int(v_text)
        except ValueError:
            return ""
        if 0 <= idx < len(shared_strings):
            return shared_strings[idx]
        return ""
    if cell_type == "b":
        if v_text is None:
            return ""
        return "TRUE" if v_text == "1" else "FALSE"
    if v_text is None:
        return ""
    return v_text


def _xlsx_sheet_to_markdown(
    zf: zipfile.ZipFile,
    sheet_path: str,
    shared_strings: list[str],
    *,
    max_rows: int,
    max_cols: int,
) -> tuple[str, dict[str, Any]]:
    data = zf.read(sheet_path)
    max_row = 0
    max_col = 0
    row_cache: list[dict[int, object | None]] = []
    with io.BytesIO(data) as handle:
        for _, elem in ET.iterparse(handle, events=("end",)):
            if _xlsx_strip_ns(elem.tag) != "row":
                continue
            row_idx_raw = elem.attrib.get("r")
            row_idx = int(row_idx_raw) if row_idx_raw and row_idx_raw.isdigit() else 0
            if row_idx == 0:
                row_idx = max_row + 1
            max_row = max(max_row, row_idx)
            row_values: dict[int, object | None] = {}
            for cell in elem:
                if _xlsx_strip_ns(cell.tag) != "c":
                    continue
                ref = cell.attrib.get("r")
                if not ref:
                    continue
                match = re.match(r"([A-Z]+)", ref)
                if not match:
                    continue
                col_idx = _excel_col_index(match.group(1))
                if col_idx <= 0:
                    continue
                max_col = max(max_col, col_idx)
                if row_idx <= max_rows and col_idx <= max_cols:
                    row_values[col_idx] = _xlsx_cell_value(cell, shared_strings)
            if row_idx <= max_rows:
                row_cache.append(row_values)
            elem.clear()
    if max_row == 0 or max_col == 0:
        return _sheet_to_markdown_table(
            [],
            total_rows=max_row,
            total_cols=max_col,
            max_rows=max_rows,
            max_cols=max_cols,
        )
    capped_cols = min(max_col, max_cols)
    rows: list[list[object | None]] = []
    for row_values in row_cache:
        row = [row_values.get(idx) for idx in range(1, capped_cols + 1)]
        rows.append(row)
    return _sheet_to_markdown_table(
        rows,
        total_rows=max_row,
        total_cols=max_col,
        max_rows=max_rows,
        max_cols=max_cols,
    )


def _apply_size_limit(
    text: str,
    max_chars: int,
    *,
    offset: int = 0,
    length: int | None = None,
) -> tuple[str, bool, dict[str, int | None]]:
    total_chars = len(text)
    safe_offset = max(0, int(offset))
    if safe_offset > total_chars:
        safe_offset = total_chars
    limit = max_chars
    if length is not None:
        limit = int(length)
    if limit <= 0:
        end = total_chars
    else:
        end = min(safe_offset + limit, total_chars)
    returned = end - safe_offset
    truncated = end < total_chars
    segment = text[safe_offset:end]
    if truncated and safe_offset == 0 and length is None and max_chars > 0:
        segment = segment.rstrip() + "\n...[truncated]..."
    paging = {
        "offset": safe_offset,
        "returned": returned,
        "total_chars": total_chars,
        "next_offset": end if end < total_chars else None,
    }
    return segment, truncated, paging


def _format_from_suffix(path: Path) -> str:
    ext = path.suffix.lower().lstrip(".")
    return ext or "text"


def _detect_format(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in {".md", ".markdown"}:
        return "md"
    if ext == ".docx":
        return "docx"
    if ext == ".pdf":
        return "pdf"
    if ext == ".xlsx":
        return "xlsx"
    return ""


def _ensure_pandoc_available() -> None:
    if _package_mode() == "full":
        pandoc_path = _resolve_pandoc_binary()
        if not pandoc_path:
            raise RuntimeError(
                "Full package mode requires a bundled pandoc binary under "
                "`dogent/resources/tools/pandoc/<platform>/`."
            )
        os.environ["PYPANDOC_PANDOC"] = str(pandoc_path)
        return
    if shutil.which("pandoc"):
        return
    try:
        import pypandoc

        pypandoc.get_pandoc_version()
        return
    except Exception as exc:
        log_exception("document_io", exc)
        raise RuntimeError(
            "Pandoc is required for DOCX export and read. "
            "Please install pandoc and try again."
        ) from exc


_HTML_IMG_TAG = re.compile(r"<img\b[^>]*>", re.IGNORECASE)
_HTML_DIV_IMG_TAG = re.compile(
    r"<div\b(?P<div_attrs>[^>]*)>\s*(?P<img_tag><img\b[^>]*>)\s*</div>",
    re.IGNORECASE | re.DOTALL,
)
_HTML_ATTR_PATTERN = re.compile(
    r"([a-zA-Z_:][\w:.-]*)(?:\s*=\s*(\"[^\"]*\"|'[^']*'|[^\s\"'>]+))?"
)


def _normalize_markdown_for_docx(md_text: str) -> tuple[str, list[str]]:
    warnings: list[str] = []

    def replace_div(match: re.Match[str]) -> str:
        div_attrs = _parse_html_attrs(match.group("div_attrs") or "")
        img_tag = match.group("img_tag") or ""
        rendered = _render_markdown_image(img_tag, div_attrs, warnings)
        return rendered if rendered else match.group(0)

    text = _HTML_DIV_IMG_TAG.sub(replace_div, md_text)

    def replace_img(match: re.Match[str]) -> str:
        rendered = _render_markdown_image(match.group(0), None, warnings)
        return rendered if rendered else match.group(0)

    text = _HTML_IMG_TAG.sub(replace_img, text)
    return text, warnings


def _render_markdown_image(
    img_tag: str,
    extra_attrs: dict[str, str] | None,
    warnings: list[str],
) -> str | None:
    attrs = _parse_html_attrs(img_tag)
    if extra_attrs:
        for key, value in extra_attrs.items():
            attrs.setdefault(key, value)
    src = attrs.pop("src", None)
    if not src:
        return None
    src = _normalize_image_src(src)
    if src is None:
        warnings.append("Skipping non-local image reference in DOCX export.")
        return None
    alt = attrs.pop("alt", "") or ""
    title = attrs.pop("title", "")
    style = attrs.pop("style", "")
    width_attr = attrs.get("width")
    height_attr = attrs.get("height")
    width_style, height_style, style_rest = _extract_style_dimensions(style)
    if not width_attr and width_style:
        attrs["width"] = width_style
    if not height_attr and height_style:
        attrs["height"] = height_style
    if style_rest:
        attrs["style"] = style_rest
    target = _format_markdown_target(src)
    title_block = f' "{title}"' if title else ""
    attrs_block = _format_pandoc_attrs(attrs)
    return f"![{alt}]({target}{title_block}){attrs_block}"


def _parse_html_attrs(raw: str) -> dict[str, str]:
    if not raw:
        return {}
    text = raw.strip()
    if text.startswith("<"):
        text = re.sub(r"^<\w+\b", "", text, flags=re.IGNORECASE).strip()
        if text.endswith(">"):
            text = text[:-1].strip()
    attrs: dict[str, str] = {}
    for match in _HTML_ATTR_PATTERN.finditer(text):
        key = match.group(1)
        raw_value = match.group(2)
        if not key:
            continue
        if raw_value is None:
            continue
        value = raw_value.strip()
        if (value.startswith("\"") and value.endswith("\"")) or (
            value.startswith("'") and value.endswith("'")
        ):
            value = value[1:-1]
        attrs[key] = value
    return attrs


def _extract_style_dimensions(style: str) -> tuple[str | None, str | None, str]:
    if not style:
        return None, None, ""
    width = None
    height = None
    remaining: list[str] = []
    for part in style.split(";"):
        item = part.strip()
        if not item:
            continue
        if ":" not in item:
            remaining.append(item)
            continue
        key, value = item.split(":", 1)
        key = key.strip().lower()
        value = value.strip()
        if key == "width" and value:
            width = value
            continue
        if key == "height" and value:
            height = value
            continue
        remaining.append(f"{key}: {value}")
    return width, height, "; ".join(remaining).strip()


def _normalize_image_src(src: str) -> str | None:
    cleaned = src.strip()
    lowered = cleaned.lower()
    if lowered.startswith(("http://", "https://", "data:")):
        return None
    if lowered.startswith("file://"):
        parsed = urlparse(cleaned)
        file_path = unquote(parsed.path or "")
        if file_path.startswith("/") and re.match(r"^/[a-zA-Z]:/", file_path):
            file_path = file_path[1:]
        return file_path or None
    if re.match(r"^[a-zA-Z]:[\\\\/]", cleaned):
        return cleaned
    if re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*:", cleaned):
        return None
    return cleaned


def _format_markdown_target(src: str) -> str:
    parsed = urlparse(src)
    if parsed.scheme:
        return src
    raw_path = unquote(parsed.path)
    if Path(raw_path).is_absolute():
        return raw_path
    target = src
    if re.search(r"[\s()<>]", target):
        return f"<{target}>"
    return target


def _format_pandoc_attrs(attrs: dict[str, str]) -> str:
    if not attrs:
        return ""
    parts: list[str] = []
    for key, value in attrs.items():
        if value == "":
            parts.append(key)
            continue
        escaped = value.replace("\"", "\\\"")
        parts.append(f'{key}="{escaped}"')
    return "{" + " ".join(parts) + "}"


def _markdown_to_docx(
    md_path: Path, *, output_path: Path, workspace_root: Path | None = None
) -> None:
    _ensure_pandoc_available()
    import pypandoc

    md_text = md_path.read_text(encoding="utf-8", errors="replace")
    normalized, _ = _normalize_markdown_for_docx(md_text)
    resource_paths = [md_path.parent.resolve()]
    if workspace_root:
        root_resolved = workspace_root.resolve()
        if root_resolved not in resource_paths:
            resource_paths.append(root_resolved)
    resource_arg = f"--resource-path={os.pathsep.join(str(p) for p in resource_paths)}"
    extra_args = ["--standalone", resource_arg, "--highlight-style=tango"]
    with tempfile.TemporaryDirectory() as tmp:
        tmp_md = Path(tmp) / md_path.name
        tmp_md.write_text(normalized, encoding="utf-8")
        pypandoc.convert_file(
            str(tmp_md),
            to="docx",
            format=(
                "markdown+raw_html+link_attributes+pipe_tables"
                "+multiline_tables+grid_tables+fenced_code_blocks"
            ),
            outputfile=str(output_path),
            extra_args=extra_args,
        )


def _default_pdf_css() -> str:
    return read_config_text(PDF_STYLE_FILENAME)


def _ensure_page_break_css(css_text: str) -> str:
    if ".page-break" in css_text:
        return css_text
    rule = ".page-break { break-after: page; page-break-after: always; }\n"
    if not css_text.strip():
        return rule
    return f"{css_text.rstrip()}\n\n{rule}"


def _resolve_pdf_style(
    workspace_root: Path | None,
    *,
    global_root: Path | None = None,
) -> tuple[str, list[str]]:
    warnings: list[str] = []
    if workspace_root:
        workspace_style = workspace_root / ".dogent" / PDF_STYLE_FILENAME
        if workspace_style.exists():
            try:
                return workspace_style.read_text(encoding="utf-8"), warnings
            except Exception as exc:
                log_exception("document_io", exc)
                warnings.append(
                    f"Could not read PDF style file: {workspace_style}. Using fallback."
                )
    resolved_global = (global_root or (Path.home() / ".dogent")) / PDF_STYLE_FILENAME
    if resolved_global.exists():
        try:
            return resolved_global.read_text(encoding="utf-8"), warnings
        except Exception as exc:
            log_exception("document_io", exc)
            warnings.append(
                f"Could not read PDF style file: {resolved_global}. Using fallback."
            )
    return _default_pdf_css(), warnings


def _build_pdf_header_footer(css_text: str) -> tuple[str | None, str | None]:
    style = f"<style>{css_text}</style>"
    footer = (
        style
        + "<div class=\"pdf-footer\">"
        + "<span class=\"pageNumber\"></span> / <span class=\"totalPages\"></span>"
        + "</div>"
    )
    return None, footer


def _markdown_to_html(
    md_text: str,
    *,
    title: str,
    css_text: str | None = None,
    base_path: Path | None = None,
    workspace_root: Path | None = None,
) -> str:
    from markdown_it import MarkdownIt

    mdi = MarkdownIt(
        "commonmark",
        {
            "html": True,
            "linkify": True,
            "typographer": True,
        },
    ).enable("table").enable("strikethrough").enable("fence")
    try:
        from pygments import highlight  # type: ignore
        from pygments.formatters import HtmlFormatter  # type: ignore
        from pygments.lexers import TextLexer, get_lexer_by_name  # type: ignore

        formatter = HtmlFormatter(nowrap=True)

        def highlight_code(code: str, lang: str, _attrs: object | None = None) -> str:
            try:
                lexer = get_lexer_by_name(lang) if lang else TextLexer()
            except Exception as exc:
                log_exception("document_io", exc)
                lexer = TextLexer()
            return highlight(code, lexer, formatter)

        mdi.options["highlight"] = highlight_code
    except Exception as exc:
        log_exception("document_io", exc)
        pass
    body = mdi.render(md_text)
    if base_path:
        body = _inline_local_images(body, base_path, workspace_root)
    css = css_text if css_text is not None else _default_pdf_css()
    css = _ensure_page_break_css(css)
    base_tag = ""
    if base_path is not None:
        base_href = base_path.resolve().as_uri()
        if not base_href.endswith("/"):
            base_href = f"{base_href}/"
        base_tag = f"<base href=\"{base_href}\" />\n"
    return (
        "<!doctype html>\n"
        "<html>\n<head>\n"
        f"<meta charset=\"utf-8\" />\n<title>{title}</title>\n"
        f"{base_tag}"
        f"<style>{css}</style>\n"
        "</head>\n<body>\n"
        f"{body}\n"
        "</body>\n</html>"
    )


def _inline_local_images(
    html: str,
    base_dir: Path,
    workspace_root: Path | None,
) -> str:
    img_pattern = re.compile(
        r'(<img\b[^>]*\bsrc=)(["\']?)([^"\'>\s]+)\2',
        re.IGNORECASE,
    )

    def replace(match: re.Match[str]) -> str:
        prefix, quote, src = match.groups()
        new_src = _to_data_uri(src, base_dir, workspace_root)
        if not new_src:
            return match.group(0)
        q = quote or '"'
        return f"{prefix}{q}{new_src}{q}"

    return img_pattern.sub(replace, html)


def _to_data_uri(
    src: str,
    base_dir: Path,
    workspace_root: Path | None,
) -> str | None:
    parsed = urlparse(src)
    if parsed.scheme in {"http", "https", "data", "file"}:
        return None
    raw_path = unquote(parsed.path)
    path = Path(raw_path)
    if not path.is_absolute():
        path = (base_dir / path).resolve()
    else:
        path = path.resolve()
    if workspace_root:
        try:
            path.relative_to(workspace_root.resolve())
        except Exception as exc:
            log_exception("document_io", exc)
            return None
    if not path.exists() or not path.is_file():
        return None
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"




async def _markdown_to_pdf(
    md_path: Path,
    *,
    output_path: Path,
    title: str | None,
    workspace_root: Path | None = None,
) -> list[str]:
    md_text = md_path.read_text(encoding="utf-8", errors="replace")
    css_text, warnings = _resolve_pdf_style(workspace_root)
    header_template, footer_template = _build_pdf_header_footer(css_text)
    html = _markdown_to_html(
        md_text,
        title=title or "Document",
        css_text=css_text,
        base_path=md_path.parent,
        workspace_root=workspace_root,
    )
    with tempfile.TemporaryDirectory() as tmp:
        html_path = Path(tmp) / "document.html"
        html_path.write_text(html, encoding="utf-8")
        await _html_to_pdf(
            html,
            output_path=output_path,
            header_template=header_template,
            footer_template=footer_template,
            source_url=html_path.resolve().as_uri(),
        )
    return warnings


async def _html_to_pdf(
    html: str,
    *,
    output_path: Path,
    header_template: str | None = None,
    footer_template: str | None = None,
    source_url: str | None = None,
) -> None:
    _configure_playwright_browsers()
    try:
        from playwright.async_api import async_playwright
    except Exception as exc:  # noqa: BLE001
        log_exception("document_io", exc)
        raise RuntimeError("PDF export requires Playwright. Install dependency.") from exc
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            if source_url:
                await page.goto(source_url, wait_until="load")
            else:
                await page.set_content(html, wait_until="load")
            pdf_options = {
                "path": str(output_path),
                "format": "A4",
                "print_background": True,
            }
            if header_template or footer_template:
                pdf_options.update(
                    {
                        "display_header_footer": True,
                        "header_template": header_template or "<span></span>",
                        "footer_template": footer_template or "<span></span>",
                        "margin": {
                            "top": "18mm",
                            "bottom": "18mm",
                            "left": "18mm",
                            "right": "18mm",
                        },
                    }
                )
            await page.pdf(**pdf_options)
            await browser.close()
    except Exception as exc:
        log_exception("document_io", exc)
        raise RuntimeError(
            "PDF export requires Playwright Chromium. Please install dependencies."
        ) from exc


def _configure_playwright_browsers() -> None:
    if _package_mode() != "full":
        return
    browsers_path = _resolve_playwright_browsers_path()
    if not browsers_path:
        raise RuntimeError(
            "Full package mode requires bundled Playwright Chromium under "
            "`dogent/resources/tools/playwright/<platform>/`."
        )
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(browsers_path)


def _ensure_playwright_chromium_installed() -> None:
    if _package_mode() == "full":
        _configure_playwright_browsers()
        return
    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except FileNotFoundError as exc:
        log_exception("document_io", exc)
        raise RuntimeError(
            "Playwright is not installed. Install dependencies to enable PDF export."
        ) from exc
    except subprocess.CalledProcessError as exc:
        log_exception("document_io", exc)
        raise RuntimeError(
            "Failed to install Chromium via Playwright. "
            f"stdout:\n{exc.stdout}\n\nstderr:\n{exc.stderr}"
        ) from exc


async def _ensure_playwright_chromium_installed_async() -> None:
    await asyncio.to_thread(_ensure_playwright_chromium_installed)


def _run_async(coro: asyncio.Future | asyncio.Task) -> Any:
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError as exc:
        log_exception("document_io", exc)
        return asyncio.run(coro)
    if loop.is_running():
        raise RuntimeError("Cannot run async export from a running event loop.")
    return loop.run_until_complete(coro)
