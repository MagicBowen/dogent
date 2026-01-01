from __future__ import annotations

import asyncio
import base64
import mimetypes
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime
from importlib import resources
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import unquote, urlparse


DEFAULT_MAX_CHARS = 15000
DEFAULT_XLSX_MAX_ROWS = 50
DEFAULT_XLSX_MAX_COLS = 20
PDF_STYLE_FILENAME = "pdf_style.css"


@dataclass(frozen=True)
class DocumentReadResult:
    content: str
    truncated: bool
    format: str
    metadata: dict[str, Any]
    error: str | None = None


@dataclass(frozen=True)
class DocumentConvertResult:
    input_format: str
    output_format: str
    output_path: Path
    extracted_media_dir: Path | None
    notes: list[str]


def read_document(
    path: Path,
    *,
    sheet: str | None = None,
    max_chars: int = DEFAULT_MAX_CHARS,
) -> DocumentReadResult:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return _read_pdf(path, max_chars=max_chars)
    if ext == ".docx":
        return _read_docx(path, max_chars=max_chars)
    if ext == ".xlsx":
        return _read_xlsx(path, sheet=sheet, max_chars=max_chars)
    return _read_text(path, max_chars=max_chars)


def export_markdown(
    md_path: Path,
    *,
    output_path: Path,
    format: str,
    title: str | None = None,
    workspace_root: Path | None = None,
) -> list[str]:
    normalized = format.strip().lower()
    if normalized == "docx":
        _markdown_to_docx(md_path, output_path=output_path)
        return []
    if normalized == "pdf":
        return _run_async(
            _markdown_to_pdf(
                md_path,
                output_path=output_path,
                title=title,
                workspace_root=workspace_root,
            )
        )
    raise ValueError(f"Unsupported export format: {format}")


async def export_markdown_async(
    md_path: Path,
    *,
    output_path: Path,
    format: str,
    title: str | None = None,
    workspace_root: Path | None = None,
) -> list[str]:
    normalized = format.strip().lower()
    if normalized == "docx":
        _markdown_to_docx(md_path, output_path=output_path)
        return []
    if normalized == "pdf":
        return await _markdown_to_pdf(
            md_path,
            output_path=output_path,
            title=title,
            workspace_root=workspace_root,
        )
    raise ValueError(f"Unsupported export format: {format}")


async def convert_document_async(
    input_path: Path,
    *,
    output_path: Path,
    extract_media_dir: Path | None = None,
    workspace_root: Path | None = None,
) -> DocumentConvertResult:
    input_format = _detect_format(input_path)
    output_format = _detect_format(output_path)
    if not input_format:
        raise ValueError("Unsupported input format. Use docx, pdf, or md.")
    if not output_format:
        raise ValueError("Unsupported output format. Use docx, pdf, or md.")
    if input_format == output_format:
        raise ValueError("Input and output formats are the same.")
    if extract_media_dir and not (
        input_format == "docx" and output_format == "md"
    ):
        raise ValueError("extract_media_dir is only supported for DOCX to Markdown.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    notes: list[str] = []

    if input_format == "docx" and output_format == "md":
        _docx_to_markdown(
            input_path,
            output_path=output_path,
            extract_media_dir=extract_media_dir,
        )
    elif input_format == "md" and output_format == "docx":
        _markdown_to_docx(input_path, output_path=output_path)
    elif input_format == "md" and output_format == "pdf":
        style_notes = await _markdown_to_pdf(
            input_path,
            output_path=output_path,
            title=output_path.stem,
            workspace_root=workspace_root,
        )
        notes.extend(style_notes)
    elif input_format == "docx" and output_format == "pdf":
        notes.append("Converted DOCX -> Markdown -> PDF; formatting may differ.")
        with tempfile.TemporaryDirectory() as tmp:
            tmp_md = Path(tmp) / "source.md"
            _docx_to_markdown(input_path, output_path=tmp_md, extract_media_dir=None)
            style_notes = await _markdown_to_pdf(
                tmp_md,
                output_path=output_path,
                title=output_path.stem,
                workspace_root=workspace_root,
            )
            notes.extend(style_notes)
    elif input_format == "pdf" and output_format == "md":
        notes.append("PDF text extracted; layout and images are not preserved.")
        result = read_document(input_path, max_chars=0)
        if result.error:
            raise RuntimeError(result.error)
        output_path.write_text(result.content, encoding="utf-8")
    elif input_format == "pdf" and output_format == "docx":
        notes.append("PDF text extracted; layout and images are not preserved.")
        result = read_document(input_path, max_chars=0)
        if result.error:
            raise RuntimeError(result.error)
        with tempfile.TemporaryDirectory() as tmp:
            tmp_md = Path(tmp) / "source.md"
            tmp_md.write_text(result.content, encoding="utf-8")
            _markdown_to_docx(tmp_md, output_path=output_path)
    else:
        raise ValueError(f"Unsupported conversion: {input_format} -> {output_format}")

    return DocumentConvertResult(
        input_format=input_format,
        output_format=output_format,
        output_path=output_path,
        extracted_media_dir=extract_media_dir,
        notes=notes,
    )


def _read_text(path: Path, *, max_chars: int) -> DocumentReadResult:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:  # noqa: BLE001
        return DocumentReadResult(
            content="",
            truncated=False,
            format=_format_from_suffix(path),
            metadata={},
            error=str(exc),
        )
    content, truncated = _apply_size_limit(text, max_chars)
    return DocumentReadResult(
        content=content,
        truncated=truncated,
        format=_format_from_suffix(path),
        metadata={},
    )


def _read_docx(path: Path, *, max_chars: int) -> DocumentReadResult:
    try:
        _ensure_pandoc_available()
        import pypandoc

        text = pypandoc.convert_file(
            str(path),
            to="markdown",
            format="docx",
            extra_args=["--track-changes=all"],
        )
    except Exception as exc:  # noqa: BLE001
        return DocumentReadResult(
            content="",
            truncated=False,
            format="docx",
            metadata={},
            error=f"DOCX read failed: {exc}",
        )
    content, truncated = _apply_size_limit(text, max_chars)
    return DocumentReadResult(
        content=content,
        truncated=truncated,
        format="docx",
        metadata={},
    )


def _docx_to_markdown(
    input_path: Path,
    *,
    output_path: Path,
    extract_media_dir: Path | None,
) -> None:
    _ensure_pandoc_available()
    import pypandoc

    extra_args = ["--track-changes=all"]
    if extract_media_dir:
        extract_media_dir.mkdir(parents=True, exist_ok=True)
        extra_args.append(f"--extract-media={extract_media_dir}")
    pypandoc.convert_file(
        str(input_path),
        to="markdown",
        format="docx",
        outputfile=str(output_path),
        extra_args=extra_args,
    )


def _read_pdf(path: Path, *, max_chars: int) -> DocumentReadResult:
    try:
        import fitz
    except Exception as exc:  # noqa: BLE001
        return DocumentReadResult(
            content="",
            truncated=False,
            format="pdf",
            metadata={},
            error=f"PDF read failed (missing PyMuPDF): {exc}",
        )

    try:
        doc = fitz.open(str(path))
    except Exception as exc:  # noqa: BLE001
        return DocumentReadResult(
            content="",
            truncated=False,
            format="pdf",
            metadata={},
            error=f"PDF read failed: {exc}",
        )

    try:
        parts: list[str] = []
        pages_with_text = 0
        for idx, page in enumerate(doc, start=1):
            text = page.get_text("text").strip()
            if text:
                pages_with_text += 1
                parts.append(f"<!-- page:{idx} -->\n\n{text}\n")
        metadata = {"pages": len(doc)}
        if pages_with_text == 0:
            return DocumentReadResult(
                content="",
                truncated=False,
                format="pdf",
                metadata=metadata,
                error="Unsupported PDF: no extractable text (scanned PDF not supported).",
            )
        combined = "\n\n".join(parts).strip() + "\n"
        content, truncated = _apply_size_limit(combined, max_chars)
        return DocumentReadResult(
            content=content,
            truncated=truncated,
            format="pdf",
            metadata=metadata,
        )
    finally:
        doc.close()


def _read_xlsx(
    path: Path,
    *,
    sheet: str | None,
    max_chars: int,
) -> DocumentReadResult:
    try:
        import openpyxl
    except Exception as exc:  # noqa: BLE001
        return DocumentReadResult(
            content="",
            truncated=False,
            format="xlsx",
            metadata={},
            error=f"XLSX read failed (missing openpyxl): {exc}",
        )
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:  # noqa: BLE001
        return DocumentReadResult(
            content="",
            truncated=False,
            format="xlsx",
            metadata={},
            error=f"XLSX read failed: {exc}",
        )
    try:
        sheet_name = _select_sheet(workbook.sheetnames, sheet)
        if not sheet_name:
            return DocumentReadResult(
                content="",
                truncated=False,
                format="xlsx",
                metadata={"sheets": workbook.sheetnames},
                error=f"Sheet not found: {sheet}",
            )
        ws = workbook[sheet_name]
        total_rows = int(ws.max_row or 0)
        total_cols = int(ws.max_column or 0)
        if total_rows == 0 or total_cols == 0:
            table, meta = "(empty sheet)", {"rows": total_rows, "cols": total_cols}
        else:
            capped_rows = min(total_rows, DEFAULT_XLSX_MAX_ROWS)
            capped_cols = min(total_cols, DEFAULT_XLSX_MAX_COLS)
            table, meta = _sheet_to_markdown_table(
                ws.iter_rows(
                    min_row=1,
                    max_row=capped_rows,
                    max_col=capped_cols,
                    values_only=True,
                ),
                total_rows=total_rows,
                total_cols=total_cols,
                max_rows=DEFAULT_XLSX_MAX_ROWS,
                max_cols=DEFAULT_XLSX_MAX_COLS,
            )
        metadata = {"sheet": sheet_name, **meta}
        content, truncated = _apply_size_limit(table, max_chars)
        return DocumentReadResult(
            content=content,
            truncated=truncated,
            format="xlsx",
            metadata=metadata,
        )
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _sheet_to_markdown_table(
    rows: Iterable[Iterable[object | None]],
    *,
    total_rows: int,
    total_cols: int,
    max_rows: int,
    max_cols: int,
) -> tuple[str, dict[str, Any]]:
    if total_rows == 0 or total_cols == 0:
        return "(empty sheet)", {"rows": total_rows, "cols": total_cols}

    capped_rows = min(total_rows, max_rows)
    capped_cols = min(total_cols, max_cols)
    raw_rows = [list(row) for row in rows]
    if not raw_rows:
        return "(empty sheet)", {"rows": total_rows, "cols": total_cols}
    header = raw_rows[0][:capped_cols]
    body_rows = raw_rows[1:capped_rows]

    if not any(_cell_to_str(cell) for cell in header):
        header = [_excel_col_name(idx + 1) for idx in range(capped_cols)]

    lines = []
    lines.append("| " + " | ".join(_cell_to_str(cell) for cell in header) + " |")
    lines.append("| " + " | ".join(["---"] * capped_cols) + " |")
    for row in body_rows:
        padded = list(row[:capped_cols])
        if len(padded) < capped_cols:
            padded.extend([None] * (capped_cols - len(padded)))
        lines.append("| " + " | ".join(_cell_to_str(cell) for cell in padded) + " |")

    truncated = total_rows > capped_rows or total_cols > capped_cols
    if truncated:
        lines.append(
            f"\n[Truncated to {capped_rows} rows x {capped_cols} cols]"
        )
    metadata = {"rows": total_rows, "cols": total_cols, "truncated": truncated}
    return "\n".join(lines), metadata


def _cell_to_str(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return str(value)


def _excel_col_name(index: int) -> str:
    name = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


def _select_sheet(sheetnames: list[str], requested: str | None) -> str | None:
    if not sheetnames:
        return None
    if requested is None or not requested.strip():
        return sheetnames[0]
    requested = requested.strip()
    if requested in sheetnames:
        return requested
    lowered = requested.lower()
    for name in sheetnames:
        if name.lower() == lowered:
            return name
    return None


def _apply_size_limit(text: str, max_chars: int) -> tuple[str, bool]:
    if max_chars <= 0 or len(text) <= max_chars:
        return text, False
    clipped = text[:max_chars].rstrip()
    return clipped + "\n...[truncated]...", True


def _format_from_suffix(path: Path) -> str:
    ext = path.suffix.lower().lstrip(".")
    return ext or "text"


def _detect_format(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in {".md", ".markdown"}:
        return "md"
    if ext == ".docx":
        return "docx"
    if ext == ".pdf":
        return "pdf"
    return ""


def _ensure_pandoc_available() -> None:
    if shutil.which("pandoc"):
        return
    try:
        import pypandoc

        pypandoc.get_pandoc_version()
        return
    except Exception:
        pass
    try:
        import pypandoc

        pypandoc.download_pandoc()
        pypandoc.get_pandoc_version()
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            "Pandoc is required for DOCX export and read. "
            "Auto-download failed. Please check network/proxy or install pandoc."
        ) from exc


def _markdown_to_docx(md_path: Path, *, output_path: Path) -> None:
    _ensure_pandoc_available()
    import pypandoc

    pypandoc.convert_file(
        str(md_path),
        to="docx",
        format="md",
        outputfile=str(output_path),
        extra_args=["--standalone"],
    )


def _default_pdf_css() -> str:
    try:
        data = resources.files("dogent").joinpath("templates").joinpath(PDF_STYLE_FILENAME)
        return data.read_text(encoding="utf-8")
    except Exception:
        return ""


def _resolve_pdf_style(
    workspace_root: Path | None,
    *,
    global_root: Path | None = None,
) -> tuple[str, list[str]]:
    warnings: list[str] = []
    if workspace_root:
        workspace_style = workspace_root / ".dogent" / PDF_STYLE_FILENAME
        if workspace_style.exists():
            try:
                return workspace_style.read_text(encoding="utf-8"), warnings
            except Exception:
                warnings.append(
                    f"Could not read PDF style file: {workspace_style}. Using fallback."
                )
    resolved_global = (global_root or (Path.home() / ".dogent")) / PDF_STYLE_FILENAME
    if resolved_global.exists():
        try:
            return resolved_global.read_text(encoding="utf-8"), warnings
        except Exception:
            warnings.append(
                f"Could not read PDF style file: {resolved_global}. Using fallback."
            )
    return _default_pdf_css(), warnings


def _build_pdf_header_footer(css_text: str) -> tuple[str | None, str | None]:
    style = f"<style>{css_text}</style>"
    footer = (
        style
        + "<div class=\"pdf-footer\">"
        + "<span class=\"pageNumber\"></span> / <span class=\"totalPages\"></span>"
        + "</div>"
    )
    return None, None


def _markdown_to_html(
    md_text: str,
    *,
    title: str,
    css_text: str | None = None,
    base_path: Path | None = None,
    workspace_root: Path | None = None,
) -> str:
    from markdown_it import MarkdownIt

    mdi = MarkdownIt(
        "commonmark",
        {
            "html": True,
            "linkify": True,
            "typographer": True,
        },
    ).enable("table").enable("strikethrough").enable("fence")
    try:
        from pygments import highlight  # type: ignore
        from pygments.formatters import HtmlFormatter  # type: ignore
        from pygments.lexers import TextLexer, get_lexer_by_name  # type: ignore

        formatter = HtmlFormatter(nowrap=True)

        def highlight_code(code: str, lang: str, _attrs: object | None = None) -> str:
            try:
                lexer = get_lexer_by_name(lang) if lang else TextLexer()
            except Exception:
                lexer = TextLexer()
            return highlight(code, lexer, formatter)

        mdi.options["highlight"] = highlight_code
    except Exception:
        pass
    body = mdi.render(md_text)
    if base_path:
        body = _inline_local_images(body, base_path, workspace_root)
    css = css_text if css_text is not None else _default_pdf_css()
    base_tag = ""
    if base_path is not None:
        base_href = base_path.resolve().as_uri()
        if not base_href.endswith("/"):
            base_href = f"{base_href}/"
        base_tag = f"<base href=\"{base_href}\" />\n"
    return (
        "<!doctype html>\n"
        "<html>\n<head>\n"
        f"<meta charset=\"utf-8\" />\n<title>{title}</title>\n"
        f"{base_tag}"
        f"<style>{css}</style>\n"
        "</head>\n<body>\n"
        f"{body}\n"
        "</body>\n</html>"
    )


def _inline_local_images(
    html: str,
    base_dir: Path,
    workspace_root: Path | None,
) -> str:
    img_pattern = re.compile(
        r'(<img\b[^>]*\bsrc=)(["\']?)([^"\'>\s]+)\2',
        re.IGNORECASE,
    )

    def replace(match: re.Match[str]) -> str:
        prefix, quote, src = match.groups()
        new_src = _to_data_uri(src, base_dir, workspace_root)
        if not new_src:
            return match.group(0)
        q = quote or '"'
        return f"{prefix}{q}{new_src}{q}"

    return img_pattern.sub(replace, html)


def _to_data_uri(
    src: str,
    base_dir: Path,
    workspace_root: Path | None,
) -> str | None:
    parsed = urlparse(src)
    if parsed.scheme in {"http", "https", "data", "file"}:
        return None
    raw_path = unquote(parsed.path)
    path = Path(raw_path)
    if not path.is_absolute():
        path = (base_dir / path).resolve()
    else:
        path = path.resolve()
    if workspace_root:
        try:
            path.relative_to(workspace_root.resolve())
        except Exception:
            return None
    if not path.exists() or not path.is_file():
        return None
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime};base64,{encoded}"




async def _markdown_to_pdf(
    md_path: Path,
    *,
    output_path: Path,
    title: str | None,
    workspace_root: Path | None = None,
) -> list[str]:
    md_text = md_path.read_text(encoding="utf-8", errors="replace")
    css_text, warnings = _resolve_pdf_style(workspace_root)
    header_template, footer_template = _build_pdf_header_footer(css_text)
    html = _markdown_to_html(
        md_text,
        title=title or "Document",
        css_text=css_text,
        base_path=md_path.parent,
        workspace_root=workspace_root,
    )
    with tempfile.TemporaryDirectory() as tmp:
        html_path = Path(tmp) / "document.html"
        html_path.write_text(html, encoding="utf-8")
        await _html_to_pdf(
            html,
            output_path=output_path,
            header_template=header_template,
            footer_template=footer_template,
            source_url=html_path.resolve().as_uri(),
        )
    return warnings


async def _html_to_pdf(
    html: str,
    *,
    output_path: Path,
    header_template: str | None = None,
    footer_template: str | None = None,
    source_url: str | None = None,
) -> None:
    try:
        from playwright.async_api import async_playwright
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("PDF export requires Playwright. Install dependency.") from exc
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            if source_url:
                await page.goto(source_url, wait_until="load")
            else:
                await page.set_content(html, wait_until="load")
            pdf_options = {
                "path": str(output_path),
                "format": "A4",
                "print_background": True,
            }
            if header_template or footer_template:
                pdf_options.update(
                    {
                        "display_header_footer": True,
                        "header_template": header_template or "<span></span>",
                        "footer_template": footer_template or "<span></span>",
                        "margin": {
                            "top": "18mm",
                            "bottom": "18mm",
                            "left": "18mm",
                            "right": "18mm",
                        },
                    }
                )
            await page.pdf(**pdf_options)
            await browser.close()
    except Exception:
        await _ensure_playwright_chromium_installed_async()
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            if source_url:
                await page.goto(source_url, wait_until="load")
            else:
                await page.set_content(html, wait_until="load")
            pdf_options = {
                "path": str(output_path),
                "format": "A4",
                "print_background": True,
            }
            if header_template or footer_template:
                pdf_options.update(
                    {
                        "display_header_footer": True,
                        "header_template": header_template or "<span></span>",
                        "footer_template": footer_template or "<span></span>",
                        "margin": {
                            "top": "18mm",
                            "bottom": "18mm",
                            "left": "18mm",
                            "right": "18mm",
                        },
                    }
                )
            await page.pdf(**pdf_options)
            await browser.close()


def _ensure_playwright_chromium_installed() -> None:
    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except FileNotFoundError as exc:
        raise RuntimeError(
            "Playwright is not installed. Install dependencies to enable PDF export."
        ) from exc
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(
            "Failed to install Chromium via Playwright. "
            f"stdout:\n{exc.stdout}\n\nstderr:\n{exc.stderr}"
        ) from exc


async def _ensure_playwright_chromium_installed_async() -> None:
    await asyncio.to_thread(_ensure_playwright_chromium_installed)


def _run_async(coro: asyncio.Future | asyncio.Task) -> Any:
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return asyncio.run(coro)
    if loop.is_running():
        raise RuntimeError("Cannot run async export from a running event loop.")
    return loop.run_until_complete(coro)
